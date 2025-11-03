from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os


class ExcelReport:
    def __init__(self, filename):
        # Ensure directory exists
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        self.filename = filename

        # Workbook setup
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Link Results"

        # Header
        self.ws.append(["Timestamp", "URL", "Status", "HTTP Code"])

        # Counters
        self.total = 0
        self.pass_count = 0
        self.fail_count = 0
        self.skip_count = 0

        # Header style (Blue)
        header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for cell in self.ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.font = Font(bold=True) 
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row
        self.ws.freeze_panes = "A2" 
    def add_row(self, url, status, code=None):
        self.total += 1
        if status.startswith("PASS"):
            self.pass_count += 1
        elif status == "FAIL":
            self.fail_count += 1
        elif status == "SKIP":
            self.skip_count += 1

        # Colors only for FAIL and SKIP (no green for PASS)
        fill_colors = {
            "FAIL": "FF7F7F",  # Light red
            "SKIP": "FFF2CC",  # Light yellow
        }
        fill = None
        if status in fill_colors:
            fill = PatternFill(start_color=fill_colors[status], end_color=fill_colors[status], fill_type="solid")

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [timestamp, url, status, code]
        self.ws.append(row)

        # Apply background color if applicable
        if fill:
            for cell in self.ws[self.ws.max_row]:
                cell.fill = fill

    def save(self):
        # Empty line before summary
        self.ws.append([])

        # Summary row
        summary_row = [
            "",
            "Summary",
            f"Total: {self.total}",
            f"Pass: {self.pass_count}, Fail: {self.fail_count}, Skip: {self.skip_count}",
            "",
        ]
        self.ws.append(summary_row)

        # Grey background + bold text
        summary_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for cell in self.ws[self.ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto column width
        for col in self.ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            self.ws.column_dimensions[col_letter].width = max_length + 2

        self.wb.save(self.filename)
